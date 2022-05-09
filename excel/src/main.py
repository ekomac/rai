from utils import (A, B, C, D, E, F, G, H, I, J, K, L, M, N, O)
from typing import Union
import openpyxl
import consts
from pregunta import Pregunta


def fill_categories_from_excel(
    file_path: str,
    categories: dict
) -> None:
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['DATOS']
    for row in sheet.rows:
        pregunta = Pregunta(row[0].value, row[1].value,
                            row[2].value, row[3].value)
        categories[pregunta.categoria].append(pregunta)


def get_value_calc_formula(
    subtipo_row,
    curr_row,
    pregunta,
    i_respuesta
) -> Union[str, int]:
    if pregunta.tipo == consts.TYPE_VEC:
        return '=IF(E{row}="+",{mas},IF(E{row}="-",{menos},0))'.format(
            row=subtipo_row,
            mas=100/pregunta.cant_respuestas*(i_respuesta+1),
            menos=100/pregunta.cant_respuestas *
            (pregunta.cant_respuestas - i_respuesta),
        )
    elif pregunta.tipo == consts.TYPE_VEX:
        return '=IF(E{row}="+",{mas},IF(E{row}="-",{menos},0))'.format(
            row=subtipo_row,
            mas=100/pregunta.cant_respuestas*(i_respuesta+1),
            menos=100/pregunta.cant_respuestas *
            (pregunta.cant_respuestas - i_respuesta),
        )
    elif pregunta.tipo == consts.TYPE_VEM:
        return '=IF(E{row}="+",{mas},IF(E{row}="-",{menos},0))'.format(
            row=subtipo_row,
            mas=100/pregunta.cant_respuestas*(i_respuesta+1),
            menos=100/pregunta.cant_respuestas *
            (pregunta.cant_respuestas - i_respuesta)
        )
    elif pregunta.tipo == consts.TYPE_BOL:
        return ('=IF(OR(AND(E{0}="+", G{1}="Si"),'
                ' AND(E{0}="-", G{1}="No")), 100, 0)').format(
            subtipo_row, curr_row)
    elif pregunta.cant_respuestas > 0:
        val = 100 / pregunta.cant_respuestas
        return '=IF(D{}="MULT",{},IF(D{}="SING",100,0))'.format(
            subtipo_row, val, subtipo_row
        )
    return 0


def set_titles(sheet) -> None:
    for row, col, title in consts.TITLES:
        sheet.cell(row=row, column=col).value = title


def merge_title_cells(sheet) -> None:
    for col in consts.COLUMNS_TO_MERGE:
        sheet.merge_cells(col)


def set_base_data(sheet, row, pregunta):
    data = [(B, pregunta.id), (C, pregunta.pregunta),
            (D, pregunta.tipo), (F, 1)]
    for col, val in data:
        sheet.cell(row=row, column=col).value = val


def merge_cell_when_various_answers(
    sheet,
    next_row,
    cant_rtas
) -> None:
    if cant_rtas > 0:
        last_row = cant_rtas + next_row-1
        for col in 'BCDEF':
            sheet.merge_cells(
                f"{col}{next_row}:{col}{last_row}")
            sheet.merge_cells(
                f"K{next_row}:K{cant_rtas + next_row-1}")
            # Combinar las celdas de los valores indexados por pregunta
            sheet.merge_cells(
                f"J{next_row}:J{cant_rtas + next_row-1}")


def apply_formatting(sheet) -> None:
    for i, row in enumerate(sheet.rows):
        for j, cell in enumerate(row):
            cell.border = consts.THIN_BORDER
            if i in [0, 1]:
                cell.font = consts.FONT
                cell.alignment = consts.TITLE_ALIGNMENT
            elif j == 2:
                cell.alignment = consts.WRAPPED_ALIGNMENT
            elif j == 7:
                cell.alignment = consts.RTA_ALIGNMENT
            else:
                cell.alignment = consts.ALIGNMENT


def apply_sum_formula(
    sheet,
    row,
    first_row,
    next_row
) -> None:
    _max = f'MAX(I{first_row}:I{next_row-1})'
    val = f'=IF(D{first_row}="MULT",SUM(I{first_row}:I{next_row-1}),{_max})'
    sheet.cell(row=row, column=J).value = val


def apply_join_formulas(
    sheet,
    row,
    first_row,
    next_row
) -> None:
    val = f'=TEXTJOIN("; ",FALSE,M{first_row}:M{next_row-1})'
    sheet.cell(row=row, column=N).value = val
    val = f'=TEXTJOIN("; "&CHAR(10),FALSE,M{first_row}:M{next_row-1})'
    sheet.cell(row=row, column=O).value = val


def add_to_summary_sheet(
    summary_sheet,
    pregunta
) -> None:
    row = len(list(summary_sheet.rows)) + 1
    summary_sheet.cell(row=row, column=A).value = pregunta.categoria
    summary_sheet.cell(row=row, column=B).value = pregunta.id
    summary_sheet.cell(row=row, column=C).value = pregunta.pregunta
    vlookup = f'=VLOOKUP(B{row},INDIRECT("\'"&A{row}&"\'' + \
        '!B3:O100"),14,FALSE)'
    summary_sheet.cell(row=row, column=D).value = vlookup


def create_new_excel(file_path: str, categories) -> None:

    # Create the excel file
    wb = openpyxl.Workbook()

    # Create the summary sheet, which will hold the final data
    wb.create_sheet("Resumen")

    # Declare de summary sheet
    summary_sheet = wb["Resumen"]

    # Set the titles for summary sheet
    for row, col, title in consts.RESUMEN_TITLES:
        summary_sheet.cell(row=row, column=col).value = title

    # Freeze first row for summary sheet
    summary_sheet.freeze_panes = "A2"

    # For each category key
    for key in categories:

        # Create a sheet for the category, with
        # the category name as the sheet name
        wb.create_sheet(key)

        # Declare the sheet
        sheet = wb[key]

        # Freeze the first and second rows
        sheet.freeze_panes = 'A3'

        # Set the titles
        set_titles(sheet)
        # Merge the title cells
        merge_title_cells(sheet)

        # First row is number 3, cause title use first and second
        curr_row = 3
        # Holds the row number of the last row that
        # will be used to the current category
        last_row = 3
        for pregunta in categories[key]:
            last_row += pregunta.cant_respuestas
        last_row -= 1

        # Set the base data for each pregunta
        for pregunta in categories[key]:

            # Sets the id, question, type and subtype for each pregunta
            set_base_data(sheet, curr_row, pregunta)

            # Paint cell when tipo is "undefined" (consts.UNDEFINED)
            if pregunta.tipo == consts.TYPE_UNDEFINED:
                sheet.cell(row=curr_row, column=D).fill = consts.UNDEFINED_FILL

            # Merge the cells when the pregunta has several answers
            merge_cell_when_various_answers(
                sheet, curr_row, pregunta.cant_respuestas)

            # Resets first_row to match next_row, being the first_row
            # the first row that will be used to the current pregunta
            first_row = curr_row

            # For each answer of the pregunta
            for i, respuesta in enumerate(pregunta.respuestas_as_list):

                # Set the answer data
                sheet.cell(row=curr_row, column=A).value = pregunta.categoria
                sheet.cell(row=curr_row, column=G).value = respuesta
                sheet.cell(row=curr_row, column=E).value = "null"
                value_cell = sheet.cell(row=curr_row, column=H)
                # Get autocalculated value for each subtipo
                value = get_value_calc_formula(
                    first_row, curr_row, pregunta, i)
                value_cell.value = value
                con_op = f'=H{curr_row}*F{first_row}'
                sheet.cell(row=curr_row, column=I).value = con_op
                final_val = '=ROUND(I%s*(100/SUM(J3:J100)),2)' % curr_row
                sheet.cell(row=curr_row, column=L).value = final_val

                # Create formula for joining the answer and it's value
                rta_and_value = f'=TEXTJOIN("=",TRUE,G{curr_row},L{curr_row})'
                # Set the formula to the cell
                sheet.cell(row=curr_row, column=M).value = rta_and_value

                # Increase row
                curr_row += 1

            apply_sum_formula(sheet, first_row, first_row, curr_row)
            apply_join_formulas(sheet, first_row, first_row, curr_row)
            if pregunta.cant_respuestas > 0:
                sheet.merge_cells(
                    f"N{first_row}:N{first_row + pregunta.cant_respuestas -1}")
                sheet.merge_cells(
                    f"O{first_row}:O{first_row + pregunta.cant_respuestas -1}")
            to_format = '=IF(D{}="MULT",{},1)'
            sheet.cell(row=first_row, column=K).value = to_format.format(
                first_row, pregunta.cant_respuestas)
            add_to_summary_sheet(summary_sheet, pregunta)

        apply_formatting(sheet)

    del wb["Sheet"]
    wb.save(file_path)


def main() -> None:
    categories = {key: [] for key in consts.CATEGORIES}
    fill_categories_from_excel('data.xlsx', categories)
    done = False
    count = 0
    while not done:
        try:
            create_new_excel(
                f'parametros_de_evaluacion-{count}.xlsx', categories)
            done = True
        except PermissionError:
            count += 1
    print("the end")


if __name__ == '__main__':
    main()
